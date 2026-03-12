from __future__ import annotations

import io
from datetime import datetime

from flask import Blueprint, Response, session, redirect, url_for, flash
from functools import wraps

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from database import db

export_finance_excel = Blueprint("export_finance_excel", __name__)


def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "user_id" not in session:
            flash("Пожалуйста, войдите в систему", "error")
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated


# ── Style helpers ─────────────────────────────────────────────────────────────

DEEP   = "1E4A54"
ACCENT = "2AA7A1"
GREEN  = "27AE60"
RED    = "C0392B"
LIGHT  = "EAF4F4"
WHITE  = "FFFFFF"


def _side():
    return Side(style="thin", color="CCCCCC")


def _border():
    s = _side()
    return Border(left=s, right=s, top=s, bottom=s)


def _fill(hex_color):
    return PatternFill("solid", start_color=hex_color, fgColor=hex_color)


def _write_title(ws, title, ncols, row=1):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=title)
    c.font = Font(name="Arial", bold=True, size=14, color=WHITE)
    c.fill = _fill(DEEP)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 30


def _write_subtitle(ws, text, ncols, row=2):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name="Arial", size=10, color="555555")
    c.fill = _fill("D6EEEE")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 18


def _write_headers(ws, headers, row, color):
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = Font(name="Arial", bold=True, size=10, color=WHITE)
        c.fill = _fill(color)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border()
    ws.row_dimensions[row].height = 22


def _write_row(ws, row, values, amount_cols=None, bg=None):
    amount_cols = amount_cols or set()
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.font = Font(name="Arial", size=10)
        c.border = _border()
        c.alignment = Alignment(vertical="center")
        if col in amount_cols:
            c.number_format = '#,##0.00'
            c.alignment = Alignment(horizontal="right", vertical="center")
        if bg:
            c.fill = _fill(bg)


def _zebra(i):
    return LIGHT if i % 2 == 0 else WHITE


def _v(obj, key):
    if isinstance(obj, dict):
        return obj.get(key, 0) or 0
    return getattr(obj, key, 0) or 0


def _payment(pt):
    return {"cash": "Наличные", "account": "Безнал"}.get(pt, pt or "")


def _date(obj, key):
    val = _v(obj, key) if isinstance(obj, dict) else getattr(obj, key, "")
    return str(val)[:16] if val else ""


# ── Sheet builders ────────────────────────────────────────────────────────────

def _summary_sheet(wb, summary, balance):
    ws = wb.create_sheet("Сводка")
    now = datetime.now().strftime("%d.%m.%Y %H:%M")
    _write_title(ws, "Финансовая сводка — Швейная Фабрика", 3)
    _write_subtitle(ws, f"Сформировано: {now}", 3, row=2)

    metrics = [
        ("Всего доходов",     _v(summary, "total_income"),   GREEN),
        ("Всего расходов",    _v(summary, "total_expenses"), RED),
        ("Прибыль",           _v(summary, "profit"),         ACCENT),
        ("Баланс (наличные)", _v(balance, "cash"),           DEEP),
        ("Баланс (безнал)",   _v(balance, "account"),        DEEP),
        ("Итоговый баланс",   _v(balance, "total"),          DEEP),
    ]

    _write_headers(ws, ["Показатель", "Сумма (сум)", ""], 4, ACCENT)

    for i, (label, value, color) in enumerate(metrics, 5):
        lc = ws.cell(row=i, column=1, value=label)
        lc.font = Font(name="Arial", bold=True, size=11)
        lc.border = _border()
        lc.fill = _fill(LIGHT)

        vc = ws.cell(row=i, column=2, value=value)
        vc.number_format = '#,##0.00'
        vc.font = Font(name="Arial", bold=True, size=12, color=color)
        vc.alignment = Alignment(horizontal="right", vertical="center")
        vc.border = _border()
        ws.row_dimensions[i].height = 24

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 8


def _expenses_sheet(wb, expenses):
    ws = wb.create_sheet("Расходы")
    headers = ["#", "Дата", "Тип расхода", "Сумма (сум)", "Оплата", "Заказ", "Комментарий", "Сотрудник"]
    ncols = len(headers)
    _write_title(ws, "Расходы", ncols)
    _write_subtitle(ws, f"Всего записей: {len(expenses)}", ncols, row=2)
    _write_headers(ws, headers, row=3, color=RED)

    for i, e in enumerate(expenses, 1):
        row = i + 3
        vals = [
            i,
            _date(e, "created_at"),
            _v(e, "expense_type"),
            _v(e, "amount"),
            _payment(_v(e, "payment_type")),
            _v(e, "order_ref") or "",
            _v(e, "comment") or "",
            _v(e, "username") or "",
        ]
        _write_row(ws, row, vals, amount_cols={4}, bg=_zebra(i))

    total_row = len(expenses) + 4
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=3)
    c = ws.cell(total_row, 1, "ИТОГО")
    c.font = Font(name="Arial", bold=True, size=11, color=WHITE)
    c.fill = _fill(RED)
    c.alignment = Alignment(horizontal="right")
    c.border = _border()
    tc = ws.cell(total_row, 4, f"=SUM(D4:D{total_row-1})")
    tc.number_format = '#,##0.00'
    tc.font = Font(name="Arial", bold=True, size=11, color=WHITE)
    tc.fill = _fill(RED)
    tc.border = _border()

    for col, w in zip("ABCDEFGH", [5, 18, 24, 18, 12, 12, 30, 18]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A4"


def _income_sheet(wb, income):
    ws = wb.create_sheet("Приходы")
    headers = ["#", "Дата", "Источник", "Сумма (сум)", "Оплата", "Комментарий", "Сотрудник"]
    ncols = len(headers)
    _write_title(ws, "Приходы", ncols)
    _write_subtitle(ws, f"Всего записей: {len(income)}", ncols, row=2)
    _write_headers(ws, headers, row=3, color=GREEN)

    for i, inc in enumerate(income, 1):
        row = i + 3
        vals = [
            i,
            _date(inc, "created_at"),
            _v(inc, "source"),
            _v(inc, "amount"),
            _payment(_v(inc, "payment_type")),
            _v(inc, "comment") or "",
            _v(inc, "username") or "",
        ]
        _write_row(ws, row, vals, amount_cols={4}, bg=_zebra(i))

    total_row = len(income) + 4
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=3)
    c = ws.cell(total_row, 1, "ИТОГО")
    c.font = Font(name="Arial", bold=True, size=11, color=WHITE)
    c.fill = _fill(GREEN)
    c.alignment = Alignment(horizontal="right")
    c.border = _border()
    tc = ws.cell(total_row, 4, f"=SUM(D4:D{total_row-1})")
    tc.number_format = '#,##0.00'
    tc.font = Font(name="Arial", bold=True, size=11, color=WHITE)
    tc.fill = _fill(GREEN)
    tc.border = _border()

    for col, w in zip("ABCDEFG", [5, 18, 24, 18, 12, 30, 18]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A4"


# ── Route ─────────────────────────────────────────────────────────────────────

@export_finance_excel.route("/api/finance/export/excel")
@login_required
def finance_export_excel():
    expenses = db.get_all_expenses(limit=5000)
    income   = db.get_all_income(limit=5000)
    balance  = db.get_cash_balance()
    summary  = db.get_financial_summary()

    wb = Workbook()
    wb.remove(wb.active)

    _summary_sheet(wb, summary, balance)
    _expenses_sheet(wb, expenses)
    _income_sheet(wb, income)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"finance_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return Response(
        buf.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
